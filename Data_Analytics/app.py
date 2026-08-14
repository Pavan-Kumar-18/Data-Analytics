from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# CREATE WORKBOOK
# =========================

workbook = Workbook()

sheet = workbook.active
sheet.title = "Sales Data"

dashboard = workbook.create_sheet("Dashboard")


# =========================
# SALES DATA
# =========================

sales = [
    {"product": "Laptop", "quantity": 2, "price": 50000},
    {"product": "Mouse", "quantity": 5, "price": 500},
    {"product": "Keyboard", "quantity": 3, "price": 1000},
    {"product": "Monitor", "quantity": 2, "price": 15000},
    {"product": "Mobile", "quantity": 4, "price": 25000},
    {"product": "Tablet", "quantity": 3, "price": 18000},
    {"product": "Headphones", "quantity": 6, "price": 2000},
    {"product": "Webcam", "quantity": 4, "price": 3500},
    {"product": "Printer", "quantity": 2, "price": 12000},
    {"product": "Scanner", "quantity": 3, "price": 8000},
    {"product": "Speaker", "quantity": 5, "price": 4500},
    {"product": "Smartwatch", "quantity": 4, "price": 7000},
    {"product": "Power Bank", "quantity": 8, "price": 1500},
    {"product": "USB Cable", "quantity": 10, "price": 400},
    {"product": "Charger", "quantity": 7, "price": 1200},
    {"product": "SSD", "quantity": 3, "price": 6500},
    {"product": "Hard Disk", "quantity": 4, "price": 5500},
    {"product": "RAM", "quantity": 5, "price": 3200},
    {"product": "Graphics Card", "quantity": 2, "price": 45000},
    {"product": "Motherboard", "quantity": 2, "price": 18000},
    {"product": "Processor", "quantity": 3, "price": 22000},
    {"product": "Cabinet", "quantity": 4, "price": 5000},
    {"product": "Keyboard Pad", "quantity": 6, "price": 700},
    {"product": "Gaming Mouse", "quantity": 5, "price": 2500},
    {"product": "Gaming Keyboard", "quantity": 3, "price": 4500},
    {"product": "Gaming Chair", "quantity": 2, "price": 15000},
    {"product": "Desk Lamp", "quantity": 5, "price": 1800},
    {"product": "Monitor Stand", "quantity": 4, "price": 2500},
    {"product": "Laptop Stand", "quantity": 6, "price": 1800},
    {"product": "Cooling Pad", "quantity": 5, "price": 2200},
    {"product": "Microphone", "quantity": 3, "price": 6000},
    {"product": "Router", "quantity": 4, "price": 3500},
    {"product": "WiFi Adapter", "quantity": 7, "price": 1200},
    {"product": "Bluetooth Adapter", "quantity": 6, "price": 800},
    {"product": "Smart TV", "quantity": 2, "price": 55000},
    {"product": "Projector", "quantity": 2, "price": 30000},
    {"product": "Projector Screen", "quantity": 3, "price": 7000},
    {"product": "HDMI Cable", "quantity": 9, "price": 600},
    {"product": "Display Cable", "quantity": 6, "price": 900},
    {"product": "Power Strip", "quantity": 8, "price": 1100},
    {"product": "UPS", "quantity": 3, "price": 8000},
    {"product": "External DVD Drive", "quantity": 2, "price": 2500},
    {"product": "Memory Card", "quantity": 10, "price": 900},
    {"product": "Card Reader", "quantity": 7, "price": 500},
    {"product": "Smartphone Case", "quantity": 12, "price": 600},
    {"product": "Screen Protector", "quantity": 15, "price": 300},
    {"product": "Earbuds", "quantity": 8, "price": 2500},
    {"product": "VR Headset", "quantity": 2, "price": 35000},
    {"product": "Game Controller", "quantity": 4, "price": 4500}
]


# =========================
# SALES DATA SHEET
# =========================

headers = ["Product", "Quantity", "Price", "Revenue"]

sheet.append(headers)


# =========================
# ANALYTICS VARIABLES
# =========================

grand_total = 0

best_selling_product = ""
best_selling_revenue = 0

best_quantity = 0
best_quantity_product = ""


# =========================
# PROCESS SALES
# =========================

for item in sales:

    product = item["product"]
    quantity = item["quantity"]
    price = item["price"]

    revenue = quantity * price

    grand_total += revenue

    # Add data to Excel
    sheet.append([
        product,
        quantity,
        price,
        revenue
    ])

    # Best product by revenue
    if revenue > best_selling_revenue:
        best_selling_revenue = revenue
        best_selling_product = product

    # Best product by quantity
    if quantity > best_quantity:
        best_quantity = quantity
        best_quantity_product = product


# Average revenue

average_revenue = grand_total / len(sales)


# =========================
# SALES TABLE STYLING
# =========================

header_fill = PatternFill(
    "solid",
    fgColor="1F4E78"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

thin_border = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2")
)


# Header styling

for cell in sheet[1]:

    cell.fill = header_fill

    cell.font = header_font

    cell.alignment = Alignment(
        horizontal="center"
    )

    cell.border = thin_border


# Data styling

for row in sheet.iter_rows(
    min_row=2,
    max_row=sheet.max_row
):

    for cell in row:

        cell.border = thin_border

        cell.alignment = Alignment(
            horizontal="center"
        )


# Currency formatting

for row in range(2, sheet.max_row + 1):

    sheet.cell(row, 3).number_format = '₹#,##0'

    sheet.cell(row, 4).number_format = '₹#,##0'


# Column widths

for column in range(1, 5):

    sheet.column_dimensions[
        get_column_letter(column)
    ].width = 18


# Freeze header

sheet.freeze_panes = "A2"


# Enable filtering

sheet.auto_filter.ref = f"A1:D{sheet.max_row}"


# =========================
# DASHBOARD TITLE
# =========================

dashboard.merge_cells("A1:H1")

dashboard["A1"] = "SALES ANALYTICS DASHBOARD"

dashboard["A1"].font = Font(
    size=20,
    bold=True,
    color="FFFFFF"
)

dashboard["A1"].fill = PatternFill(
    "solid",
    fgColor="17365D"
)

dashboard["A1"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

dashboard.row_dimensions[1].height = 35


# =========================
# KPI DATA
# =========================

kpis = [

    ("A3", "B3",
     "Total Revenue",
     grand_total),

    ("D3", "E3",
     "Average Revenue",
     average_revenue),

    ("A6", "B6",
     "Best Selling Product",
     best_selling_product),

    ("D6", "E6",
     "Best Selling Revenue",
     best_selling_revenue),

    ("A9", "B9",
     "Top Product by Quantity",
     best_quantity_product),

    ("D9", "E9",
     "Highest Quantity Sold",
     best_quantity)
]


# Different colors for KPI cards

fills = [

    "1F4E78",
    "548235",
    "7030A0",
    "C55A11",
    "2F75B5",
    "70AD47"
]


# =========================
# CREATE KPI CARDS
# =========================

for index, (
    label_cell,
    value_cell,
    label,
    value
) in enumerate(kpis):

    fill = PatternFill(
        "solid",
        fgColor=fills[index]
    )

    # Label

    dashboard[label_cell] = label

    dashboard[label_cell].font = Font(
        color="FFFFFF",
        bold=True,
        size=11
    )

    dashboard[label_cell].fill = fill

    dashboard[label_cell].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    dashboard[label_cell].border = thin_border


    # Value

    dashboard[value_cell] = value

    dashboard[value_cell].font = Font(
        color="FFFFFF",
        bold=True,
        size=14
    )

    dashboard[value_cell].fill = fill

    dashboard[value_cell].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    dashboard[value_cell].border = thin_border


# Currency formatting

dashboard["B3"].number_format = '₹#,##0'

dashboard["E3"].number_format = '₹#,##0'

dashboard["E6"].number_format = '₹#,##0'


# =========================
# PIE CHART
# =========================

pie = PieChart()

pie.title = "Revenue Distribution by Product"

pie.height = 8

pie.width = 12


# Revenue column D

revenue_data = Reference(
    sheet,
    min_col=4,
    min_row=1,
    max_row=5
)


# Product column A

product_categories = Reference(
    sheet,
    min_col=1,
    min_row=2,
    max_row=5
)


# Add revenue data

pie.add_data(
    revenue_data,
    titles_from_data=True
)


# Add product names

pie.set_categories(
    product_categories
)


# Legend

pie.legend.position = "r"


# Add pie chart to dashboard

dashboard.add_chart(
    pie,
    "A12"
)


# =========================
# BAR CHART
# =========================

bar = BarChart()

bar.type = "col"

bar.style = 10

bar.title = "Quantity Sold by Product"

bar.y_axis.title = "Quantity"

bar.x_axis.title = "Product"

bar.height = 8

bar.width = 12


# Quantity column B

quantity_data = Reference(
    sheet,
    min_col=2,
    min_row=1,
    max_row=5
)


# Add quantity data

bar.add_data(
    quantity_data,
    titles_from_data=True
)


# Add product names

bar.set_categories(
    product_categories
)


# Add bar chart to dashboard

dashboard.add_chart(
    bar,
    "H12"
)


# =========================
# SAVE FILE
# =========================

workbook.save(
    "Sales_Analytics_Dashboard.xlsx"
)


print(
    "Professional Sales Analytics Dashboard created successfully!"
)