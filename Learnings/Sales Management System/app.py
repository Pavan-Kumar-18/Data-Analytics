from openpyxl import Workbook
from openpyxl import load_workbook
import os

def add_product():
    try:
        num_product = int(input("Enter the Number of Products to add: "))
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Product Name", "Quantity", "Price", "Total"])
        grand_total = 0
        for i in range(num_product):
            total_price = 0
            product_name = input("Enter Product Name: ")
            product_quantity = int(input("Enter Qunatity: "))
            product_price = int(input("Enter Price: "))
            total_price += product_quantity * product_price
            grand_total += total_price
            sheet.append([product_name, product_quantity, product_price, total_price])
        sheet.append(["Grand Total", grand_total])
        workbook.save("sales.xlsx")
        print("Sales Report Created")
    except ValueError:
        print("Enter Proper Values")

def view_sales():
    try:
      file_path = "sales.xlsx"
      if os.path.isfile(file_path):
          workboook = load_workbook(file_path)
          sheet = workboook.active
          print("=====Sales======")
          for rows in sheet.iter_rows(values_only=True):
              data = list(rows)
              product_name = data[0]
              product_quanity = data[1]
              product_price = data[2]
              product_total = data[3]
              print(f"{product_name} | {product_quanity} | {product_price} | {product_total}")
    except ValueError:
        print("File Name Doesn't Exist")

def search_product():
    try:
        file_path = "sales.xlsx"
        if os.path.isfile(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
            search_term = input("Enter the Product Name:")
            found = False
            for rows in sheet.iter_rows(values_only=True):
                if rows[0].lower() == search_term.lower():
                    print(rows)
                    found = True
                    break
            if not found:
                print("Item Not Founded")
    except ValueError:
        print("Enter the correct Value to search")


def update_product():
    try:
        file_path = "sales.xlsx"
        if os.path.isfile(file_path):
            found = False
            workbook = load_workbook(file_path)
            sheet = workbook.active
            search_term = input("Enter the Product Name to update: ")
            for rows in sheet.iter_rows(min_row=2):
                if rows[0].value.lower() == search_term.lower():
                    print([cell.value for cell in rows])
                    update_quantity = int(input("Enter New Quantity:"))
                    update_price = int(input("Enter New Price:"))
                    rows[1].value = update_quantity
                    rows[2].value = update_price
                    rows[3].value = update_price * update_quantity
                    found = True
                    break
        if found:
            workbook.save(file_path)
            print("Data Updated")
        else:
            print("Product Not Founded")
    except ValueError:
        print("Data not founded")

def delete_product():
    try:
        file_name = "sales.xlsx"
        if os.path.isfile(file_name):
            workbook = load_workbook(file_name)
            sheet = workbook.active
            delete_data = input("Enter the product name :")
            found = False
            for rows in sheet.iter_rows(min_row=2):
                if rows[0].value.lower() == delete_data.lower():
                    sheet.delete_rows(rows[0].row, 1)
                    found = True
                    break
            if found:
                print("Data Deleted")
                workbook.save(file_name)
            else:
                print("Data Not Founded")
    except ValueError:
        print("Values Not Founded")


def sales_management():
    try:
        print("===== Sales Report =======")
        print("1.Add Product")
        print("2.View Sales")
        print("3.Search Product")
        print("4.Update Product")
        print("5.Delete Product")
        user_choice = int(input("Enter Choice:"))
        if user_choice == 1:
            add_product()
        elif user_choice == 2:
            view_sales()
        elif user_choice == 3:
            search_product()
        elif user_choice == 4:
            update_product()
        elif user_choice == 5:
            delete_product()
    except ValueError:
        print("Somthing Went Wrong")

print(sales_management())

